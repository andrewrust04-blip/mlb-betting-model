# export_excel.py
# Excel formatting and export functions for the MLB pitcher strikeout betting system.
# Called by live_predictions.py, settle_bets.py, and performance_dashboard.py.
# Does NOT modify any CSV files or pipeline logic.

import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config_github import BASE_DIR

# =============================================================================
# PATHS
# =============================================================================

FILTERED_BETS_XLSX = os.path.join(BASE_DIR, "filtered_bets.xlsx")
BET_LOG_XLSX       = os.path.join(BASE_DIR, "bet_log.xlsx")
DASHBOARD_XLSX     = os.path.join(BASE_DIR, "performance_dashboard.xlsx")

# =============================================================================
# SHARED STYLE CONSTANTS
# =============================================================================

C_HEADER_DARK = "1F2D3D"
C_HEADER_TEXT = "FFFFFF"
C_WIN         = "D6EFD8"
C_WIN_DARK    = "2D6A4F"
C_LOSS        = "FAD7D7"
C_LOSS_DARK   = "A4262C"
C_UNSETTLED   = "FFF8E7"
C_OVER        = "D6EAF8"
C_UNDER       = "EAD6F8"
C_SUBHEADER   = "2E4057"
C_POSITIVE    = "1A7431"
C_NEGATIVE    = "C0392B"
C_BORDER      = "CCCCCC"

FONT_NAME = "Arial"

def _font(bold=False, size=10, color="000000"):
    return Font(name=FONT_NAME, bold=bold, size=size, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center")

def _left():
    return Alignment(horizontal="left", vertical="center")

def _apply_header_row(ws, row_num, labels, col_start=1):
    for i, label in enumerate(labels):
        c = ws.cell(row=row_num, column=col_start + i, value=label)
        c.font = _font(bold=True, color=C_HEADER_TEXT)
        c.fill = _fill(C_HEADER_DARK)
        c.alignment = _center()
        c.border = _border()

def _apply_subheader(ws, row_num, label, n_cols, col_start=1):
    for i in range(n_cols):
        c = ws.cell(row=row_num, column=col_start + i)
        c.fill = _fill(C_SUBHEADER)
    ws.cell(row=row_num, column=col_start).value = label
    ws.cell(row=row_num, column=col_start).font = _font(bold=True, size=11, color=C_HEADER_TEXT)
    ws.cell(row=row_num, column=col_start).alignment = _left()

def _write_cell(cell, val, fmt, row_bg):
    if fmt == "date":
        cell.value = pd.to_datetime(val).date() if pd.notna(val) else ""
        cell.number_format = "MM/DD/YYYY"
    elif fmt == "pct":
        cell.value = float(val) if pd.notna(val) and val != "" else ""
        if pd.notna(val) and val != "":
            cell.number_format = "0.0%"
    elif fmt == "roi":
        cell.value = float(val) if pd.notna(val) and val != "" else ""
        if pd.notna(val) and val != "":
            cell.number_format = "0.0%"
            fv = float(val)
            cell.font = _font(bold=True,
                              color=C_POSITIVE if fv > 0 else C_NEGATIVE if fv < 0 else "000000")
    elif fmt == "units":
        cell.value = float(val) if pd.notna(val) and val != "" else ""
        if pd.notna(val) and val != "":
            cell.number_format = "+0.000;-0.000;0.000"
            fv = float(val)
            cell.font = _font(color=C_POSITIVE if fv > 0 else C_NEGATIVE if fv < 0 else "000000")
    elif fmt == "profit":
        cell.value = float(val) if pd.notna(val) and val != "" else ""
        if pd.notna(val) and val != "":
            cell.number_format = "+0.000;-0.000;0.000"
            fv = float(val)
            cell.font = _font(bold=True,
                              color=C_POSITIVE if fv > 0 else C_NEGATIVE if fv < 0 else "000000")
    elif fmt == "int":
        cell.value = int(val) if pd.notna(val) and val != "" else ""
    elif fmt == "int_or_blank":
        try:
            cell.value = int(val) if pd.notna(val) and val != "" else ""
        except Exception:
            cell.value = ""
    elif fmt == "0.0":
        cell.value = float(val) if pd.notna(val) and val != "" else ""
        if pd.notna(val) and val != "":
            cell.number_format = "0.0"
    elif fmt == "0.00":
        cell.value = float(val) if pd.notna(val) and val != "" else ""
        if pd.notna(val) and val != "":
            cell.number_format = "0.00"
    else:
        cell.value = str(val) if pd.notna(val) else ""

    if fmt not in ("roi", "units", "profit"):
        cell.font = _font()
    cell.fill = _fill(row_bg)
    cell.alignment = _center()
    cell.border = _border()


# =============================================================================
# FILTERED BETS XLSX
# =============================================================================

def export_filtered_bets(filtered_bets_df):
    """Export today's filtered bets to a clean color-coded Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Today's Bets"
    ws.sheet_view.showGridLines = False

    df = filtered_bets_df.copy().sort_values("edge", ascending=False).reset_index(drop=True)

    # Title
    ws.merge_cells("A1:K1")
    date_str = pd.to_datetime(df["date"].iloc[0]).strftime("%B %d, %Y") if len(df) > 0 else ""
    tc = ws["A1"]
    tc.value = f"MLB Strikeout Prop Bets — {date_str}"
    tc.font = _font(bold=True, size=13, color=C_HEADER_TEXT)
    tc.fill = _fill(C_HEADER_DARK)
    tc.alignment = _center()
    ws.row_dimensions[1].height = 28

    # Summary
    n_over  = int((df["bet_side"] == "over").sum())
    n_under = int((df["bet_side"] == "under").sum())
    total_rec = df["recommended_units"].sum()
    ws.merge_cells("A2:K2")
    sc = ws["A2"]
    sc.value = (f"{len(df)} bets today  |  {n_over} overs  |  {n_under} unders  "
                f"|  {total_rec:.1f} recommended units at risk")
    sc.font = _font(size=10, color="444444")
    sc.fill = _fill("EEF2F7")
    sc.alignment = _center()
    ws.row_dimensions[2].height = 18

    columns = [
        ("Pitcher",      "pitcher_name",      22, "text"),
        ("Team",         "team",               7, "text"),
        ("Opponent",     "opponent",           9, "text"),
        ("Line",         "line",               8, "0.0"),
        ("Side",         "bet_side",           8, "text"),
        ("Odds",         "odds",               8, "int"),
        ("Model Prob",   "model_prob",        11, "pct"),
        ("Implied Prob", "implied_prob",      11, "pct"),
        ("Edge",         "edge",               9, "pct"),
        ("Rec Units",    "recommended_units", 10, "0.0"),
        ("Kelly Units",  "kelly_units",       10, "0.00"),
    ]

    _apply_header_row(ws, 3, [c[0] for c in columns])
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"

    for i, (_, _, width, _) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for r_idx, row in df.iterrows():
        er = r_idx + 4
        is_over = str(row.get("bet_side", "")).lower() == "over"
        bg = (C_OVER if is_over else C_UNDER) if r_idx % 2 == 0 else ("E8F4FB" if is_over else "F3EBF9")

        for col_idx, (_, col_key, _, fmt) in enumerate(columns, start=1):
            cell = ws.cell(row=er, column=col_idx)
            _write_cell(cell, row.get(col_key, ""), fmt, bg)
            if col_key == "pitcher_name":
                cell.alignment = _left()
            if col_key == "edge" and pd.notna(row.get("edge")) and float(row.get("edge", 0)) > 0.10:
                cell.font = _font(bold=True, color=C_POSITIVE)
        ws.row_dimensions[er].height = 17

    wb.save(FILTERED_BETS_XLSX)
    print(f"Saved filtered_bets.xlsx to: {FILTERED_BETS_XLSX}")


# =============================================================================
# BET LOG XLSX
# =============================================================================

def export_bet_log(bet_log_df):
    """Export full bet log to Excel. Green = win, red = loss, amber = unsettled."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bet Log"
    ws.sheet_view.showGridLines = False

    df = bet_log_df.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df.sort_values("date", ascending=False).reset_index(drop=True)

    settled = df[df["settled"].astype(str).str.lower() == "true"].copy()
    wins         = int((settled["bet_result"] == "win").sum())
    losses       = int((settled["bet_result"] == "loss").sum())
    pushes       = int((settled["bet_result"] == "push").sum())
    flat_profit  = settled["profit_units"].sum() if len(settled) > 0 else 0.0
    rec_profit   = (settled["profit_units"] * settled["recommended_units"]).sum() if len(settled) > 0 else 0.0

    # Title
    ws.merge_cells("A1:L1")
    tc = ws["A1"]
    tc.value = "MLB Strikeout Props — Bet Log"
    tc.font = _font(bold=True, size=13, color=C_HEADER_TEXT)
    tc.fill = _fill(C_HEADER_DARK)
    tc.alignment = _center()
    ws.row_dimensions[1].height = 28

    # KPI summary rows
    kpis = [
        ("Total Bets", len(df)),
        ("Settled",    len(settled)),
        ("Wins",       wins),
        ("Losses",     losses),
        ("Pushes",     pushes),
        ("Win Rate",   f"{wins/(wins+losses):.1%}" if (wins+losses) > 0 else "—"),
        ("Flat P&L",   f"{flat_profit:+.3f} u"),
        ("Rec P&L",    f"{rec_profit:+.3f} u"),
    ]
    for i, (label, value) in enumerate(kpis):
        lc = ws.cell(row=2, column=i+1, value=label)
        lc.font = _font(bold=True, size=8, color="666666")
        lc.fill = _fill("EEF2F7")
        lc.alignment = _center()
        lc.border = _border()

        vc = ws.cell(row=3, column=i+1, value=value)
        is_pos = isinstance(value, str) and value.startswith("+")
        is_neg = isinstance(value, str) and value.startswith("-")
        vc.font = _font(bold=True, size=11,
                        color=C_POSITIVE if is_pos else C_NEGATIVE if is_neg else "000000")
        vc.fill = _fill("FFFFFF")
        vc.alignment = _center()
        vc.border = _border()

    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 8

    columns = [
        ("Date",          "date",              13, "date"),
        ("Pitcher",       "pitcher_name",       22, "text"),
        ("Team",          "team",               7, "text"),
        ("Opponent",      "opponent",           9, "text"),
        ("Line",          "line",               8, "0.0"),
        ("Side",          "bet_side",           8, "text"),
        ("Odds",          "odds",               8, "int"),
        ("Model Prob",    "model_prob",         11, "pct"),
        ("Implied Prob",  "implied_prob",       11, "pct"),
        ("Edge",          "edge",               9, "pct"),
        ("Pred Mean",     "predicted_mean",     10, "0.00"),
        ("Rec Units",     "recommended_units",  10, "0.0"),
        ("Kelly Units",   "kelly_units",        10, "0.00"),
        ("Actual K",      "actual_strikeouts",   9, "int_or_blank"),
        ("Result",        "bet_result",          8, "text"),
        ("Profit (flat)", "profit_units",        12, "profit"),
    ]

    HEADER_ROW = 5
    _apply_header_row(ws, HEADER_ROW, [c[0] for c in columns])
    ws.row_dimensions[HEADER_ROW].height = 20
    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    for i, (_, _, width, _) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for r_idx, row in df.iterrows():
        er = r_idx + HEADER_ROW + 1
        result = str(row.get("bet_result", "")).lower().strip()
        is_settled = str(row.get("settled", "")).lower() == "true"

        if result == "win":
            bg = C_WIN
        elif result == "loss":
            bg = C_LOSS
        elif not is_settled:
            bg = C_UNSETTLED
        else:
            bg = "F5F5F5"

        for col_idx, (_, col_key, _, fmt) in enumerate(columns, start=1):
            cell = ws.cell(row=er, column=col_idx)
            _write_cell(cell, row.get(col_key, ""), fmt, bg)
            if col_key == "pitcher_name":
                cell.alignment = _left()
            if col_key == "bet_result" and fmt != "profit":
                cell.font = _font(
                    bold=True,
                    color=C_WIN_DARK if result == "win" else C_LOSS_DARK if result == "loss" else "000000"
                )
        ws.row_dimensions[er].height = 17

    wb.save(BET_LOG_XLSX)
    print(f"Saved bet_log.xlsx to: {BET_LOG_XLSX}")


# =============================================================================
# PERFORMANCE DASHBOARD XLSX
# =============================================================================

def _write_breakdown(ws, start_row, title, summary_df, col_map):
    n_cols = len(col_map)
    _apply_subheader(ws, start_row, title, n_cols)
    ws.row_dimensions[start_row].height = 20
    _apply_header_row(ws, start_row + 1, [c[0] for c in col_map])
    ws.row_dimensions[start_row + 1].height = 18

    for col_idx, (_, _, width, _) in enumerate(col_map, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for r_idx, row in summary_df.reset_index(drop=True).iterrows():
        er = start_row + 2 + r_idx
        bg = "F7F9FC" if r_idx % 2 == 0 else "FFFFFF"
        for col_idx, (_, col_key, _, fmt) in enumerate(col_map, start=1):
            cell = ws.cell(row=er, column=col_idx)
            _write_cell(cell, row.get(col_key, ""), fmt, bg)
        ws.row_dimensions[er].height = 17

    return start_row + 2 + len(summary_df) + 2


def export_performance_dashboard(bet_log_df):
    """
    Build a multi-sheet performance dashboard Excel file.
    Sheets: Summary | By Line | By Edge | Over vs Under | Daily | All Bets
    """
    wb = Workbook()

    df = bet_log_df.copy()
    df["date"]              = pd.to_datetime(df["date"], errors="coerce")
    df["profit_units"]      = pd.to_numeric(df.get("profit_units"),      errors="coerce")
    df["edge"]              = pd.to_numeric(df.get("edge"),              errors="coerce")
    df["odds"]              = pd.to_numeric(df.get("odds"),              errors="coerce")
    df["line"]              = pd.to_numeric(df.get("line"),              errors="coerce")
    df["recommended_units"] = pd.to_numeric(df.get("recommended_units"), errors="coerce").fillna(1.0)
    df["kelly_units"]       = pd.to_numeric(df.get("kelly_units"),       errors="coerce").fillna(0.0)
    df["settled"]           = df["settled"].astype(str).str.lower().map({"true": True, "false": False}).fillna(False)

    settled = df[df["settled"] == True].copy()
    settled["profit_flat"]        = settled["profit_units"]
    settled["profit_recommended"] = settled["profit_units"] * settled["recommended_units"]
    settled["risk_flat"]          = 1.0
    settled["risk_recommended"]   = settled["recommended_units"]

    wins    = int((settled["bet_result"] == "win").sum())
    losses  = int((settled["bet_result"] == "loss").sum())
    pushes  = int((settled["bet_result"] == "push").sum())
    decision_bets = wins + losses

    flat_profit  = settled["profit_flat"].sum()
    rec_profit   = settled["profit_recommended"].sum()
    flat_risk    = settled["risk_flat"].sum()
    rec_risk     = settled["risk_recommended"].sum()
    flat_roi     = flat_profit / flat_risk  if flat_risk  > 0 else np.nan
    rec_roi      = rec_profit  / rec_risk   if rec_risk   > 0 else np.nan
    win_rate     = wins / decision_bets     if decision_bets > 0 else np.nan

    # ---- SHEET 1: SUMMARY ----
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:F1")
    ws1["A1"].value = "MLB Strikeout Props — Performance Dashboard"
    ws1["A1"].font = _font(bold=True, size=14, color=C_HEADER_TEXT)
    ws1["A1"].fill = _fill(C_HEADER_DARK)
    ws1["A1"].alignment = _center()
    ws1.row_dimensions[1].height = 32

    date_range = ""
    if len(settled) > 0:
        mn = settled["date"].min().strftime("%b %d")
        mx = settled["date"].max().strftime("%b %d, %Y")
        date_range = f"{mn} – {mx}"

    ws1.merge_cells("A2:F2")
    ws1["A2"].value = f"Based on {len(settled)} settled bets   |   {date_range}"
    ws1["A2"].font = _font(size=10, color="555555")
    ws1["A2"].fill = _fill("EEF2F7")
    ws1["A2"].alignment = _center()
    ws1.row_dimensions[2].height = 18

    kpi_data = [
        ("Settled Bets", len(settled),   "int"),
        ("Wins",         wins,            "int"),
        ("Losses",       losses,          "int"),
        ("Pushes",       pushes,          "int"),
        ("Win Rate",     win_rate,        "pct"),
        ("",             "",              ""),
        ("Flat Profit",  flat_profit,     "units"),
        ("Flat ROI",     flat_roi,        "pct"),
        ("Rec Profit",   rec_profit,      "units"),
        ("Rec ROI",      rec_roi,         "pct"),
    ]

    _apply_subheader(ws1, 4, "Overall Performance", 2)
    ws1.row_dimensions[4].height = 20

    for i, (label, value, fmt) in enumerate(kpi_data):
        r = 5 + i
        lc = ws1.cell(row=r, column=1, value=label)
        lc.font = _font(bold=True)
        lc.fill = _fill("F0F4FA" if i % 2 == 0 else "FFFFFF")
        lc.alignment = _left()
        lc.border = _border()

        vc = ws1.cell(row=r, column=2)
        _write_cell(vc, value, fmt, "F0F4FA" if i % 2 == 0 else "FFFFFF")
        ws1.row_dimensions[r].height = 18

    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 14

    # ---- SHEET 2: BY LINE ----
    ws2 = wb.create_sheet("By Line")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:H1")
    ws2["A1"].value = "Performance by Strikeout Line"
    ws2["A1"].font = _font(bold=True, size=13, color=C_HEADER_TEXT)
    ws2["A1"].fill = _fill(C_HEADER_DARK)
    ws2["A1"].alignment = _center()
    ws2.row_dimensions[1].height = 28

    line_summary = (
        settled.groupby("line", dropna=False)
        .agg(bets=("line","count"), wins=("bet_result", lambda x: (x=="win").sum()),
             losses=("bet_result", lambda x: (x=="loss").sum()),
             flat_profit=("profit_flat","sum"), flat_risk=("risk_flat","sum"),
             rec_profit=("profit_recommended","sum"), rec_risk=("risk_recommended","sum"))
        .reset_index().sort_values("line")
    )
    line_summary["win_rate"] = np.where((line_summary["wins"]+line_summary["losses"])>0,
        line_summary["wins"]/(line_summary["wins"]+line_summary["losses"]), np.nan)
    line_summary["flat_roi"] = np.where(line_summary["flat_risk"]>0, line_summary["flat_profit"]/line_summary["flat_risk"], np.nan)
    line_summary["rec_roi"]  = np.where(line_summary["rec_risk"]>0,  line_summary["rec_profit"]/line_summary["rec_risk"],   np.nan)

    _write_breakdown(ws2, 3, "Breakdown by Line", line_summary, [
        ("Line",     "line",        10, "0.0"),
        ("Bets",     "bets",         8, "int"),
        ("Wins",     "wins",         8, "int"),
        ("Losses",   "losses",       8, "int"),
        ("Win Rate", "win_rate",    10, "pct"),
        ("Flat P&L", "flat_profit", 12, "units"),
        ("Flat ROI", "flat_roi",    10, "roi"),
        ("Rec ROI",  "rec_roi",     10, "roi"),
    ])

    # ---- SHEET 3: BY EDGE ----
    ws3 = wb.create_sheet("By Edge")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:H1")
    ws3["A1"].value = "Performance by Edge Bucket"
    ws3["A1"].font = _font(bold=True, size=13, color=C_HEADER_TEXT)
    ws3["A1"].fill = _fill(C_HEADER_DARK)
    ws3["A1"].alignment = _center()
    ws3.row_dimensions[1].height = 28

    bins   = [0, 0.05, 0.075, 0.10, 0.15, 1.00]
    labels = ["0–5%", "5–7.5%", "7.5–10%", "10–15%", "15%+"]
    settled["edge_bucket"] = pd.cut(settled["edge"], bins=bins, labels=labels, include_lowest=True)
    edge_summary = (
        settled.groupby("edge_bucket", dropna=False, observed=False)
        .agg(bets=("edge","count"), wins=("bet_result", lambda x: (x=="win").sum()),
             losses=("bet_result", lambda x: (x=="loss").sum()),
             avg_edge=("edge","mean"), flat_profit=("profit_flat","sum"),
             flat_risk=("risk_flat","sum"), rec_profit=("profit_recommended","sum"),
             rec_risk=("risk_recommended","sum"))
        .reset_index()
    )
    edge_summary["win_rate"] = np.where((edge_summary["wins"]+edge_summary["losses"])>0,
        edge_summary["wins"]/(edge_summary["wins"]+edge_summary["losses"]), np.nan)
    edge_summary["flat_roi"] = np.where(edge_summary["flat_risk"]>0, edge_summary["flat_profit"]/edge_summary["flat_risk"], np.nan)
    edge_summary["rec_roi"]  = np.where(edge_summary["rec_risk"]>0,  edge_summary["rec_profit"]/edge_summary["rec_risk"],   np.nan)

    _write_breakdown(ws3, 3, "Breakdown by Edge Bucket", edge_summary, [
        ("Edge Bucket", "edge_bucket",  14, "text"),
        ("Bets",        "bets",          8, "int"),
        ("Wins",        "wins",          8, "int"),
        ("Losses",      "losses",        8, "int"),
        ("Avg Edge",    "avg_edge",     10, "pct"),
        ("Win Rate",    "win_rate",     10, "pct"),
        ("Flat P&L",    "flat_profit",  12, "units"),
        ("Flat ROI",    "flat_roi",     10, "roi"),
    ])

    # ---- SHEET 4: OVER vs UNDER ----
    ws4 = wb.create_sheet("Over vs Under")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:H1")
    ws4["A1"].value = "Over vs Under Performance"
    ws4["A1"].font = _font(bold=True, size=13, color=C_HEADER_TEXT)
    ws4["A1"].fill = _fill(C_HEADER_DARK)
    ws4["A1"].alignment = _center()
    ws4.row_dimensions[1].height = 28

    side_summary = (
        settled.groupby("bet_side", dropna=False)
        .agg(bets=("bet_side","count"), wins=("bet_result", lambda x: (x=="win").sum()),
             losses=("bet_result", lambda x: (x=="loss").sum()),
             flat_profit=("profit_flat","sum"), flat_risk=("risk_flat","sum"),
             rec_profit=("profit_recommended","sum"), rec_risk=("risk_recommended","sum"))
        .reset_index()
    )
    side_summary["win_rate"] = np.where((side_summary["wins"]+side_summary["losses"])>0,
        side_summary["wins"]/(side_summary["wins"]+side_summary["losses"]), np.nan)
    side_summary["flat_roi"] = np.where(side_summary["flat_risk"]>0, side_summary["flat_profit"]/side_summary["flat_risk"], np.nan)
    side_summary["rec_roi"]  = np.where(side_summary["rec_risk"]>0,  side_summary["rec_profit"]/side_summary["rec_risk"],   np.nan)

    _write_breakdown(ws4, 3, "Over vs Under Breakdown", side_summary, [
        ("Side",     "bet_side",    10, "text"),
        ("Bets",     "bets",         8, "int"),
        ("Wins",     "wins",         8, "int"),
        ("Losses",   "losses",       8, "int"),
        ("Win Rate", "win_rate",    10, "pct"),
        ("Flat P&L", "flat_profit", 12, "units"),
        ("Flat ROI", "flat_roi",    10, "roi"),
        ("Rec ROI",  "rec_roi",     10, "roi"),
    ])

    # ---- SHEET 5: DAILY ----
    ws5 = wb.create_sheet("Daily")
    ws5.sheet_view.showGridLines = False
    ws5.merge_cells("A1:H1")
    ws5["A1"].value = "Daily Results"
    ws5["A1"].font = _font(bold=True, size=13, color=C_HEADER_TEXT)
    ws5["A1"].fill = _fill(C_HEADER_DARK)
    ws5["A1"].alignment = _center()
    ws5.row_dimensions[1].height = 28

    daily_summary = (
        settled.groupby(settled["date"].dt.date)
        .agg(bets=("date","count"), wins=("bet_result", lambda x: (x=="win").sum()),
             losses=("bet_result", lambda x: (x=="loss").sum()),
             flat_profit=("profit_flat","sum"), flat_risk=("risk_flat","sum"),
             rec_profit=("profit_recommended","sum"), rec_risk=("risk_recommended","sum"))
        .reset_index().rename(columns={"date":"bet_date"}).sort_values("bet_date", ascending=False)
    )
    daily_summary["win_rate"] = np.where((daily_summary["wins"]+daily_summary["losses"])>0,
        daily_summary["wins"]/(daily_summary["wins"]+daily_summary["losses"]), np.nan)
    daily_summary["flat_roi"] = np.where(daily_summary["flat_risk"]>0, daily_summary["flat_profit"]/daily_summary["flat_risk"], np.nan)
    daily_summary["rec_roi"]  = np.where(daily_summary["rec_risk"]>0,  daily_summary["rec_profit"]/daily_summary["rec_risk"],   np.nan)

    _write_breakdown(ws5, 3, "Day-by-Day Results", daily_summary, [
        ("Date",     "bet_date",    13, "text"),
        ("Bets",     "bets",         8, "int"),
        ("Wins",     "wins",         8, "int"),
        ("Losses",   "losses",       8, "int"),
        ("Win Rate", "win_rate",    10, "pct"),
        ("Flat P&L", "flat_profit", 12, "units"),
        ("Flat ROI", "flat_roi",    10, "roi"),
        ("Rec ROI",  "rec_roi",     10, "roi"),
    ])

    # ---- SHEET 6: ALL BETS ----
    ws6 = wb.create_sheet("All Bets")
    ws6.sheet_view.showGridLines = False
    ws6.merge_cells("A1:L1")
    ws6["A1"].value = "All Settled Bets"
    ws6["A1"].font = _font(bold=True, size=13, color=C_HEADER_TEXT)
    ws6["A1"].fill = _fill(C_HEADER_DARK)
    ws6["A1"].alignment = _center()
    ws6.row_dimensions[1].height = 28

    all_cols = [
        ("Date",      "date",              13, "date"),
        ("Pitcher",   "pitcher_name",       22, "text"),
        ("Team",      "team",               7, "text"),
        ("Opponent",  "opponent",           9, "text"),
        ("Line",      "line",               8, "0.0"),
        ("Side",      "bet_side",           8, "text"),
        ("Odds",      "odds",               8, "int"),
        ("Edge",      "edge",               9, "pct"),
        ("Rec Units", "recommended_units", 10, "0.0"),
        ("Actual K",  "actual_strikeouts",  9, "int_or_blank"),
        ("Result",    "bet_result",         9, "text"),
        ("Profit",    "profit_units",       11, "profit"),
    ]

    _apply_header_row(ws6, 2, [c[0] for c in all_cols])
    ws6.row_dimensions[2].height = 20
    ws6.freeze_panes = "A3"

    for i, (_, _, width, _) in enumerate(all_cols, start=1):
        ws6.column_dimensions[get_column_letter(i)].width = width

    for r_idx, row in settled.sort_values("date", ascending=False).reset_index(drop=True).iterrows():
        er = r_idx + 3
        result = str(row.get("bet_result", "")).lower().strip()
        bg = C_WIN if result == "win" else C_LOSS if result == "loss" else "F5F5F5"

        for col_idx, (_, col_key, _, fmt) in enumerate(all_cols, start=1):
            cell = ws6.cell(row=er, column=col_idx)
            _write_cell(cell, row.get(col_key, ""), fmt, bg)
            if col_key == "pitcher_name":
                cell.alignment = _left()
            if col_key == "bet_result" and fmt != "profit":
                cell.font = _font(bold=True,
                                  color=C_WIN_DARK if result=="win" else C_LOSS_DARK if result=="loss" else "000000")
        ws6.row_dimensions[er].height = 17

    wb.save(DASHBOARD_XLSX)
    print(f"Saved performance_dashboard.xlsx to: {DASHBOARD_XLSX}")